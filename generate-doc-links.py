#!/usr/bin/env python3
# Copyright Contributors to the Open Cluster Management project

"""
Prerequisites:
- Python 3.10 or higher
- openpyxl library (pip install openpyxl)
- pytorch library (pip install torch torchvision pillow)
- playwright library (pip install playwright) (playwright install)

Usage:
  ./generate-doc-links.py [--compare]

This script will extract the DOC_LINKS constants from the doc-util.tsx file
and save them to an Excel file named doc-links.xlsx.
The Excel file will have three sheets: ACM, OCP, and Other.
The ACM sheet will have the DOC_LINKS constants for the ACM documentation.
The OCP sheet will have the DOC_LINKS constants for the OCP documentation.
The Other sheet will have the DOC_LINKS constants for the external links.

The --compare option will enable screenshot comparison using AI and generate similarity scores.
Check pass/fail status are based on the similarity scores and the thresholds defined in the script.
The screenshots are saved in the current working dir, this also acts as a continuation mechanism
where it will skip taking the screenshot of a link if a screenshot exists. It's useful for
times when your internet drops or when the VPN disconnects and you want to continue where the
script left off.

Import the doc-links.xlsx to Google Sheets
Set import location to 'Insert new sheet(s)'

Limitations:
- The Excel checkbox doesn't work in Google Sheets. So just highlight the Checked column and use insert
checkbox feature. Google sheets will automatically convert the boolean values to checkboxes.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

DOC_FILE = Path("./frontend/src/lib/doc-util.tsx")
OUTPUT_XLSX = Path("doc-links.xlsx")

DOC_VERSION_RE = re.compile(r"export const DOC_VERSION\s*=\s*['\"]([^'\"]+)['\"]")
OCP_DOC_VERSION_RE = re.compile(r"export const OCP_DOC_VERSION\s*=\s*['\"]([^'\"]+)['\"]")
DOC_HOME_RE = re.compile(r"export const DOC_HOME\s*=\s*`([^`]+)`")
OCP_DOC_HOME_RE = re.compile(r"const OCP_DOC_HOME\s*=\s*`([^`]+)`")
DOC_BASE_RE = re.compile(r"export const DOC_BASE_PATH\s*=\s*`([^`]+)`")
OCP_DOC_BASE_RE = re.compile(r"const OCP_DOC_BASE_PATH\s*=\s*`([^`]+)`")

DOC_LINKS_SECTION_RE = re.compile(r"export const DOC_LINKS = \{(.*?)\n\}", re.DOTALL)
DOC_LINKS_LINE_RE = re.compile(r"^ +([A-Z_]+):")
TEMPLATE_VAR_RE = re.compile(r"\$\{([A-Z0-9_]+)\}")
DOC_STAGING_BASE_TEMPLATE = (
    "https://content-stage.docs.redhat.com/en/documentation/"
    "red_hat_advanced_cluster_management_for_kubernetes/{DOC_VERSION}/html-single"
)
OCP_DOC_STAGING_BASE_TEMPLATE = (
    "https://content-stage.docs.redhat.com/en/documentation/"
    "openshift_container_platform/{OCP_DOC_VERSION}/html-single"
)
ACM_DOC_THRESHOLD = 0.91
OCP_DOC_THRESHOLD = 0.90

def find_doc_file() -> Path:
    if DOC_FILE.is_file():
        return DOC_FILE

    print(f"File {DOC_FILE} not found. Searching...")
    for path in Path(".").rglob("doc-util.*"):
        if path.is_file():
            return path

    print("Could not find doc-util file. Please check the path.")
    sys.exit(1)


def resolve_template(template: str, variables: dict[str, str]) -> str:
    if not template:
        return ""

    resolved = template
    for _ in range(5):
        replaced = TEMPLATE_VAR_RE.sub(
            lambda match: variables.get(match.group(1), match.group(0)),
            resolved,
        )
        if replaced == resolved:
            break
        resolved = replaced
    return resolved


def read_doc_paths(doc_text: str) -> tuple[str, str, str, str, str, str, str, str]:
    doc_base_match = DOC_BASE_RE.search(doc_text)
    ocp_doc_base_match = OCP_DOC_BASE_RE.search(doc_text)
    doc_version_match = DOC_VERSION_RE.search(doc_text)
    doc_home_match = DOC_HOME_RE.search(doc_text)
    ocp_doc_version_match = OCP_DOC_VERSION_RE.search(doc_text)
    ocp_doc_home_match = OCP_DOC_HOME_RE.search(doc_text)

    doc_base_template = doc_base_match.group(1) if doc_base_match else ""
    ocp_doc_base_template = ocp_doc_base_match.group(1) if ocp_doc_base_match else ""
    doc_version = doc_version_match.group(1) if doc_version_match else ""
    doc_home_template = doc_home_match.group(1) if doc_home_match else ""
    ocp_doc_version = ocp_doc_version_match.group(1) if ocp_doc_version_match else ""
    ocp_doc_home_template = ocp_doc_home_match.group(1) if ocp_doc_home_match else ""

    doc_home = resolve_template(doc_home_template, {"DOC_VERSION": doc_version})
    doc_base_path = resolve_template(
        doc_base_template,
        {"DOC_HOME": doc_home, "DOC_VERSION": doc_version},
    )
    ocp_doc_home = resolve_template(
        ocp_doc_home_template, {"OCP_DOC_VERSION": ocp_doc_version}
    )
    ocp_doc_base_path = resolve_template(
        ocp_doc_base_template,
        {"OCP_DOC_HOME": ocp_doc_home, "OCP_DOC_VERSION": ocp_doc_version},
    )

    print(f"Found DOC_BASE_PATH: {doc_base_path or doc_base_template}")
    print(f"Found OCP_DOC_BASE_PATH: {ocp_doc_base_path or ocp_doc_base_template}")
    print(f"Found DOC_VERSION: {doc_version}")

    return (
        doc_base_path,
        ocp_doc_base_path,
        doc_version,
        doc_home_template,
        doc_base_template,
        ocp_doc_version,
        ocp_doc_home_template,
        ocp_doc_base_template,
    )


def extract_doc_links_section(doc_text: str) -> list[str]:
    match = DOC_LINKS_SECTION_RE.search(doc_text)
    if not match:
        return []

    section = match.group(1)
    lines = section.splitlines()
    return [line for line in lines if DOC_LINKS_LINE_RE.match(line)]


def extract_relative_path(line: str, base_name: str) -> str:
    if base_name == "DOC_BASE_PATH":
        pattern = r".*\$\{DOC_BASE_PATH\}([^,`\"']+).*"
    else:
        pattern = r".*\$\{OCP_DOC_BASE_PATH\}([^,`\"']+).*"

    match = re.match(pattern, line)
    return match.group(1) if match else ""


def extract_url(line: str) -> str:
    match = re.search(r"[^\"'`]*[\"'`]([^\"'`]+)[\"'`]", line)
    return match.group(1) if match else ""


def decrement_version(version: str) -> str:
    if not version:
        return ""

    parts = version.split(".")
    if len(parts) == 1:
        try:
            value = int(parts[0])
        except ValueError:
            return ""
        return str(value - 1) if value > 0 else ""

    major, minor = parts[0], parts[1]
    try:
        minor_value = int(minor)
    except ValueError:
        return ""

    if minor_value <= 0:
        return ""

    return f"{major}.{minor_value - 1}"


def derive_doc_base_path_for_version(doc_base_path: str, version: str, target_version: str) -> str:
    if not doc_base_path or not version or not target_version:
        return ""

    version_token = f"/{version}/"
    target_token = f"/{target_version}/"
    if version_token in doc_base_path:
        return doc_base_path.replace(version_token, target_token, 1)

    if version in doc_base_path:
        return doc_base_path.replace(version, target_version, 1)

    return ""

def take_screenshot(url, name, anchor_selector):
    """Take a screenshot and return the HTTP status code (0 if unknown)."""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        if Path(f"{name}.png").is_file():
            print(f"Skipping screenshot for {url} because {name}.png already exists")
            return 0
        browser = p.chromium.launch()
        page = browser.new_page()
        response = page.goto(url, wait_until="domcontentloaded")
        status = response.status if response else 0
        if status >= 400:
            print(f"HTTP {status} for {url}")
        page.wait_for_load_state("domcontentloaded")
        # Wait for network to settle so async content loads before screenshotting.
        try:
            page.wait_for_load_state("networkidle", timeout=60000)
        except Exception:
            pass
        # Ensure the fragment anchor is applied before we manipulate the DOM
        if anchor_selector:
            try:
                page.wait_for_selector(anchor_selector, state="visible", timeout=60000)
                page.locator(anchor_selector).first.scroll_into_view_if_needed()
            except Exception:
                print(f"Anchor not found before timeout; taking screenshot anyway: {url} ({anchor_selector})")

        # Look for cookie banners and close them if present
        cookie_selectors = [
            'button:has-text("Accept all")',
            'button:has-text("Accept All")',
            'button:has-text("Accept")',
            'button:has-text("I agree")',
            'button:has-text("Agree")',
            '[aria-label="Accept all"]',
            '[aria-label="Accept cookies"]',
            '[aria-label="Close"]',
            '[data-testid="cookie-accept"]',
        ]
        for selector in cookie_selectors:
            try:
                button = page.locator(selector).first
                if button.is_visible():
                    button.click()
                    page.wait_for_timeout(500)
                    break
            except Exception:
                pass

        # Remove sticky headers and breadcrumb links that cause visual diffs
        page.add_style_tag(content="""
            header, .pf-v5-c-page__header, .pf-c-page__header,
            .masthead, .site-header, .rh-header,
            nav[aria-label="Breadcrumb"], nav[aria-label="Breadcrumbs"],
            .pf-v5-c-breadcrumb, .pf-c-breadcrumb, .breadcrumb,
            [data-testid="breadcrumbs"], .docs-breadcrumbs, .rh-breadcrumbs,
            #breadcrumbs, .breadcrumbs, .breadcrumbs-docs,
            .feedback-container-content-page {
                display: none !important;
            }
            body { margin-top: 0 !important; }
        """)
        page.evaluate("""
            () => {
                const selectors = [
                    'header',
                    '.pf-v5-c-page__header',
                    '.pf-c-page__header',
                    '.masthead',
                    '.site-header',
                    '.rh-header',
                    'nav[aria-label="Breadcrumb"]',
                    'nav[aria-label="Breadcrumbs"]',
                    '.pf-v5-c-breadcrumb',
                    '.pf-c-breadcrumb',
                    '.breadcrumb',
                    '[data-testid="breadcrumbs"]',
                    '.docs-breadcrumbs',
                    '.rh-breadcrumbs',
                    '#breadcrumbs',
                    '.breadcrumbs',
                    '.breadcrumbs-docs',
                    '.feedback-container-content-page'
                ];
                selectors.forEach((sel) => {
                    document.querySelectorAll(sel).forEach((el) => el.remove());
                });
            }
        """)

        # full_page=True captures the entire scrollable area
        page.screenshot(path=f"{name}.png", full_page=True)
        browser.close()
        return status

class SmartComparator:
    def __init__(self, crop_left_px=360):
        import torch  # noqa: F811
        from torchvision import models, transforms  # noqa: F811
        from PIL import Image  # noqa: F811

        self._torch = torch
        self._Image = Image
        self._transforms = transforms

        # Load a pre-trained VGG16 model
        vgg = models.vgg16(weights=models.VGG16_Weights.DEFAULT)
        self.feature_extractor = vgg.features

        # Set to evaluation mode (locks the model so it doesn't learn)
        self.feature_extractor.eval()

        # 2. Define the image transformation pipeline
        # Resize to standard size and normalize to match what VGG expects
        self.preprocess = self._transforms.Compose([
            self._transforms.Resize((224, 224)),
            self._transforms.ToTensor(),
            self._transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                       std=[0.229, 0.224, 0.225]),
        ])
        self.crop_left_px = crop_left_px

    def _crop_left_nav(self, img):
        if self.crop_left_px <= 0:
            return img
        width, height = img.size
        left = min(self.crop_left_px, width)
        # Crop from left edge to remove the nav, keep the content area.
        return img.crop((left, 0, width, height))

    def get_embedding(self, image_path):
        """Turn an image into a dense vector of numbers"""
        img = self._Image.open(image_path).convert('RGB')
        img = self._crop_left_nav(img)
        img_tensor = self.preprocess(img).unsqueeze(0) # Add batch dimension

        with self._torch.no_grad():
            # Pass image through the network
            features = self.feature_extractor(img_tensor)

        # Flatten the features into a 1D vector
        return features.flatten()

    def compare(self, img_path1, img_path2):
        vec1 = self.get_embedding(img_path1)
        vec2 = self.get_embedding(img_path2)

        # 3. Calculate Cosine Similarity
        # 1.0 = Identical
        # 0.0 = Completely different
        cos = self._torch.nn.CosineSimilarity(dim=0, eps=1e-6)
        similarity = cos(vec1, vec2)

        return similarity.item()

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate doc-links.xlsx from DOC_LINKS constants.",
    )
    parser.add_argument(
        "--compare",
        action="store_true",
        help="Enable screenshot comparison and similarity scores.",
    )
    args = parser.parse_args()
    enable_compare = args.compare

    doc_file = find_doc_file()
    print(f"Using file: {doc_file}")

    doc_text = doc_file.read_text(encoding="utf-8")
    (
        doc_base_path,
        ocp_doc_base_path,
        doc_version,
        doc_home_template,
        doc_base_template,
        ocp_doc_version,
        ocp_doc_home_template,
        ocp_doc_base_template,
    ) = read_doc_paths(doc_text)
    doc_version_minus1 = decrement_version(doc_version)
    doc_home_minus1 = resolve_template(
        doc_home_template,
        {"DOC_VERSION": doc_version_minus1},
    )
    doc_base_path_minus1 = (
        resolve_template(
            doc_base_template,
            {"DOC_HOME": doc_home_minus1, "DOC_VERSION": doc_version_minus1},
        )
        if doc_home_minus1
        else ""
    )
    if not doc_base_path_minus1:
        doc_base_path_minus1 = derive_doc_base_path_for_version(
            doc_base_path,
            doc_version,
            doc_version_minus1,
        )
    doc_staging_base_path = (
        DOC_STAGING_BASE_TEMPLATE.format(DOC_VERSION=doc_version) if doc_version else ""
    )
    ocp_doc_staging_base_path = (
        OCP_DOC_STAGING_BASE_TEMPLATE.format(OCP_DOC_VERSION=ocp_doc_version) if ocp_doc_version else ""
    )
    ocp_doc_version_minus1 = decrement_version(ocp_doc_version)
    ocp_doc_home_minus1 = resolve_template(
        ocp_doc_home_template,
        {"OCP_DOC_VERSION": ocp_doc_version_minus1},
    )
    ocp_doc_base_path_minus1 = (
        resolve_template(
            ocp_doc_base_template,
            {"OCP_DOC_HOME": ocp_doc_home_minus1, "OCP_DOC_VERSION": ocp_doc_version_minus1},
        )
        if ocp_doc_home_minus1
        else ""
    )
    if not ocp_doc_base_path_minus1:
        ocp_doc_base_path_minus1 = derive_doc_base_path_for_version(
            ocp_doc_base_path,
            ocp_doc_version,
            ocp_doc_version_minus1,
        )
    lines = extract_doc_links_section(doc_text)
    if not lines:
        print(f"No DOC_LINKS constants found in {doc_file}")
        sys.exit(1)

    workbook = Workbook()
    acm_sheet = workbook.active
    acm_sheet.title = "ACM"
    ocp_sheet = workbook.create_sheet("OCP")
    other_sheet = workbook.create_sheet("Other")

    version_minus1_label = doc_version_minus1 or "DOC_VERSION-1"
    version_label = doc_version or "DOC_VERSION"
    acm_sheet.append(
        [
            "Constant",
            "Relative to DOC_BASE_PATH",
            f"ACM {version_minus1_label} Link",
            f"ACM {version_label} Link",
            "Checked",
        ]
    )
    if enable_compare:
        acm_sheet["F1"] = "Similarity Score"
    ocp_version_minus1_label = ocp_doc_version_minus1 or "OCP_DOC_VERSION-1"
    ocp_version_label = ocp_doc_version or "OCP_DOC_VERSION"
    ocp_sheet.append(
        [
            "Constant",
            "Relative to OCP_DOC_BASE_PATH",
            f"OCP {ocp_version_minus1_label} Link",
            f"OCP {ocp_version_label} Link",
            "Checked",
        ]
    )
    if enable_compare:
        ocp_sheet["F1"] = "Similarity Score"
    other_sheet.append(["Constant", "URL", "Checked"])

    acm_sheet["A1"].font = Font(bold=True)
    acm_sheet["B1"].font = Font(bold=True)
    acm_sheet["C1"].font = Font(bold=True)
    acm_sheet["D1"].font = Font(bold=True)
    acm_sheet["E1"].font = Font(bold=True)
    if enable_compare:
        acm_sheet["F1"].font = Font(bold=True)
    ocp_sheet["A1"].font = Font(bold=True)
    ocp_sheet["B1"].font = Font(bold=True)
    ocp_sheet["C1"].font = Font(bold=True)
    ocp_sheet["D1"].font = Font(bold=True)
    ocp_sheet["E1"].font = Font(bold=True)
    if enable_compare:
        ocp_sheet["F1"].font = Font(bold=True)
    other_sheet["A1"].font = Font(bold=True)
    other_sheet["B1"].font = Font(bold=True)
    other_sheet["C1"].font = Font(bold=True)

    num_links_acm = 0
    num_links_ocp = 0
    num_links_other = 0

    comparator = SmartComparator(crop_left_px=360) if enable_compare else None

    for line in lines:
        key_match = DOC_LINKS_LINE_RE.match(line)
        key = key_match.group(1) if key_match else ""

        if "${DOC_BASE_PATH}" in line:
            relative_path = extract_relative_path(line, "DOC_BASE_PATH")
            row_index = acm_sheet.max_row + 1
            acm_sheet.append([key, relative_path, None, None, False])
            if doc_base_path_minus1:
                acm_sheet[
                    f"C{row_index}"
                ] = f'=HYPERLINK("{doc_base_path_minus1}"&$B{row_index}, "Link")'
            if doc_staging_base_path:
                acm_sheet[
                    f"D{row_index}"
                ] = f'=HYPERLINK("{doc_staging_base_path}"&$B{row_index}, "Link")'
            num_links_acm += 1

            if enable_compare:
                if not doc_base_path_minus1 or not doc_staging_base_path:
                    acm_sheet[f"E{row_index}"] = False
                    acm_sheet[f"F{row_index}"] = "Missing ACM compare base URL"
                    continue
                # take screenshot of the links
                print(f"Relative path: {relative_path}")
                anchor = relative_path.find("#")
                print(f"Anchor: {anchor}")
                if anchor != -1:
                    anchor = relative_path[anchor:]
                else:
                    anchor = ""
                print(f"Taking screenshot of {doc_base_path_minus1}{relative_path} with anchor {anchor}")
                status1 = take_screenshot(f"{doc_base_path_minus1}{relative_path}", f"{version_minus1_label}acm_link{row_index}", anchor)
                print(f"Taking screenshot of {doc_staging_base_path}{relative_path} with anchor {anchor}")
                status2 = take_screenshot(f"{doc_staging_base_path}{relative_path}", f"{version_label}acm_link{row_index}", anchor)

                http_error = (status1 and status1 >= 400) or (status2 and status2 >= 400)
                if http_error:
                    print(f"❌ HTTP error detected (status {status1}, {status2}) — marking as failed")
                    acm_sheet[f"E{row_index}"] = False
                    acm_sheet[f"F{row_index}"] = f"HTTP {status1}/{status2}"
                else:
                    # compare the screenshots
                    score = comparator.compare(f"{version_minus1_label}acm_link{row_index}.png", f"{version_label}acm_link{row_index}.png")
                    print(f"Similarity Score: {score:.4f}")
                    if score < ACM_DOC_THRESHOLD:
                        print("❌ ALERT: Visual Regression Detected!")
                        acm_sheet[f"E{row_index}"] = False
                    else:
                        print("✅ Visuals look stable.")
                        acm_sheet[f"E{row_index}"] = True
                    acm_sheet[f"F{row_index}"] = f"{score:.4f}"
        elif "${OCP_DOC_BASE_PATH}" in line:
            relative_path = extract_relative_path(line, "OCP_DOC_BASE_PATH")
            row_index = ocp_sheet.max_row + 1
            ocp_sheet.append([key, relative_path, None, None, False])
            if ocp_doc_base_path_minus1:
                ocp_sheet[
                    f"C{row_index}"
                ] = f'=HYPERLINK("{ocp_doc_base_path_minus1}"&$B{row_index}, "Link")'
            if ocp_doc_staging_base_path:
                ocp_sheet[
                    f"D{row_index}"
                ] = f'=HYPERLINK("{ocp_doc_staging_base_path}"&$B{row_index}, "Link")'
            num_links_ocp += 1

            if enable_compare:
                if not ocp_doc_base_path_minus1 or not ocp_doc_staging_base_path:
                    ocp_sheet[f"E{row_index}"] = False
                    ocp_sheet[f"F{row_index}"] = "Missing OCP compare base URL"
                    continue
                # take screenshot of the links
                print(f"Relative path: {relative_path}")
                anchor = relative_path.find("#")
                print(f"Anchor: {anchor}")
                if anchor != -1:
                    anchor = relative_path[anchor:]
                else:
                    anchor = ""
                print(f"Taking screenshot of {ocp_doc_base_path_minus1}{relative_path} with anchor {anchor}")
                status1 = take_screenshot(f"{ocp_doc_base_path_minus1}{relative_path}", f"{ocp_version_minus1_label}ocp_link{row_index}", anchor)
                print(f"Taking screenshot of {ocp_doc_staging_base_path}{relative_path} with anchor {anchor}")
                status2 = take_screenshot(f"{ocp_doc_staging_base_path}{relative_path}", f"{ocp_version_label}ocp_link{row_index}", anchor)

                http_error = (status1 and status1 >= 400) or (status2 and status2 >= 400)
                if http_error:
                    print(f"❌ HTTP error detected (status {status1}, {status2}) — marking as failed")
                    ocp_sheet[f"E{row_index}"] = False
                    ocp_sheet[f"F{row_index}"] = f"HTTP {status1}/{status2}"
                else:
                    # compare the screenshots
                    score = comparator.compare(f"{ocp_version_minus1_label}ocp_link{row_index}.png", f"{ocp_version_label}ocp_link{row_index}.png")
                    print(f"Similarity Score: {score:.4f}")
                    if score < OCP_DOC_THRESHOLD:
                        print("❌ ALERT: Visual Regression Detected!")
                        ocp_sheet[f"E{row_index}"] = False
                    else:
                        print("✅ Visuals look stable.")
                        ocp_sheet[f"E{row_index}"] = True
                    ocp_sheet[f"F{row_index}"] = f"{score:.4f}"
        else:
            url = extract_url(line)
            row_index = other_sheet.max_row + 1
            other_sheet.append([key, None, False])
            if url:
                other_sheet[
                    f"B{row_index}"
                ] = f'=HYPERLINK("{url}", "{url}")'
            num_links_other += 1

    if acm_sheet.max_row > 1:
        checkbox_validation = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=False,
        )
        acm_sheet.add_data_validation(checkbox_validation)
        checkbox_validation.add(f"E2:E{acm_sheet.max_row}")
    if ocp_sheet.max_row > 1:
        checkbox_validation = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=False,
        )
        ocp_sheet.add_data_validation(checkbox_validation)
        checkbox_validation.add(f"E2:E{ocp_sheet.max_row}")

    workbook.save(OUTPUT_XLSX)

    total_links = num_links_acm + num_links_ocp + num_links_other
    if total_links == 0:
        print(f"No DOC_LINKS constants found in {doc_file}")
        sys.exit(1)

    print(f"Successfully extracted {total_links} DOC_LINKS constants:")
    print(f"  - {num_links_acm} ACM documentation links")
    print(f"  - {num_links_ocp} OCP documentation links")
    print(f"  - {num_links_other} external links")
    print(f"Output saved to: {OUTPUT_XLSX}")
    print("You can now open this Excel file in your spreadsheet application.")


if __name__ == "__main__":
    main()
